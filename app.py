from threading import Timer

import psutil
import qrcode
from flask import Flask, render_template, request, jsonify, redirect, url_for, abort
import io
import base64
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import webbrowser
from collections import deque
from functools import wraps

app = Flask(__name__)

PORT = 31009
INPUT_EXCEL = 'input.xlsx'
OUTPUT_EXCEL = 'output.xlsx'

# Use a deque to store the last submissions
last_submissions = deque(maxlen=10)


def open_browser():
    webbrowser.open_new(f'http://localhost:{PORT}/private')


def get_ip_addresses():
    ip_addresses = []
    for interface, addrs in psutil.net_if_addrs().items():
        for addr in addrs:
            if addr.family == 2:  # AF_INET (IPv4)
                ip_addresses.append((interface, addr.address))
    return ip_addresses


def generate_qr_code(data):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    return img


def initialize_output_excel():
    if not os.path.exists(INPUT_EXCEL):
        raise FileNotFoundError(f"{INPUT_EXCEL} not found. Please provide the input Excel file.")

    df = pd.read_excel(INPUT_EXCEL)
    df['Attended'] = ''
    df['Submit Time'] = ''
    df['IP'] = ''
    df['User Agent'] = ''
    df.to_excel(OUTPUT_EXCEL, index=False)


def localhost_only(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if request.remote_addr != '127.0.0.1':
            abort(403)  # Forbidden
        return f(*args, **kwargs)

    return decorated_function


@app.route('/private')
@localhost_only
def index():
    ip_addresses = get_ip_addresses()
    return render_template('index.html', ip_addresses=ip_addresses, last_submissions=list(last_submissions))


@app.route('/private/generate_qr', methods=['POST'])
@localhost_only
def generate_qr():
    selected_ip = request.json.get('ip')
    full_url = f"http://{selected_ip}:{PORT}/public/attendance"
    qr_img = generate_qr_code(full_url)
    img_buffer = io.BytesIO()
    qr_img.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    qr_code = base64.b64encode(img_buffer.getvalue()).decode()
    return jsonify({'qr_code': qr_code, 'full_url': full_url})


@app.route('/public/attendance', methods=['GET', 'POST'])
def attendance():
    if request.method == 'POST':
        student_id = request.form['student_id']
        ip_address = request.remote_addr
        user_agent = request.user_agent.string
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        result = update_attendance(student_id, timestamp, ip_address, user_agent)
        if result == 'duplicate':
            return redirect(url_for('attendance_error', error='duplicate'))
        elif result == 'student_not_found':
            return redirect(url_for('attendance_error', error='not_found'))
        return redirect(url_for('attendance_success'))
    return render_template('attendance_form.html')


@app.route('/attendance_success')
def attendance_success():
    return render_template('attendance_success.html')


@app.route('/attendance_error')
def attendance_error():
    error = request.args.get('error', 'unknown')
    if error == 'duplicate':
        message = "Your attendance has been recorded, but it appears to be a duplicate submission."
    elif error == 'not_found':
        message = "Student ID not found. Please check your ID and try again."
    else:
        message = "An unknown error occurred. Please try again later."
    return render_template('attendance_error.html', message=message)


def update_attendance(student_id, timestamp, ip_address, user_agent):
    wb = load_workbook(OUTPUT_EXCEL)
    ws = wb.active

    student_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if str(row[0].value) == str(student_id):
            student_row = row
            break

    if student_row is None:
        wb.close()
        return 'student_not_found'

    row_index = student_row[0].row
    is_duplicate = False
    duplicate_with = None

    # Check for duplicate IP or User Agent
    for other_row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=6):
        if other_row[0].value == ip_address or other_row[1].value == user_agent:
            is_duplicate = True
            duplicate_with = ws.cell(row=other_row[0].row, column=1).value
            break

    # Check if this student has already attended
    if ws.cell(row=row_index, column=3).value:
        is_duplicate = True

    # Always save submission details
    ws.cell(row=row_index, column=4, value=timestamp)
    ws.cell(row=row_index, column=5, value=ip_address)
    ws.cell(row=row_index, column=6, value=user_agent)

    if is_duplicate:
        duplicate_cell = ws.cell(row=row_index, column=3)
        if duplicate_with:
            duplicate_cell.value = f"Duplicated with {duplicate_with}"
        else:
            duplicate_cell.value = "Duplicate submission"
        duplicate_cell.font = Font(color="FF0000")
        duplicate_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    else:
        # Mark attendance
        ws.cell(row=row_index, column=3, value='X')

        # Add to last 5 submissions
        student_name = ws.cell(row=row_index, column=2).value
        last_submissions.appendleft({
            'Student ID': student_id,
            'Student Name': student_name,
            'Submit Time': timestamp
        })

    wb.save(OUTPUT_EXCEL)
    wb.close()
    return 'duplicate' if is_duplicate else 'success'


@app.route('/private/get_last_submissions')
@localhost_only
def get_last_submissions():
    return jsonify(list(last_submissions))


if __name__ == '__main__':
    if not os.path.exists(OUTPUT_EXCEL):
        initialize_output_excel()
    Timer(1, open_browser).start()
    app.run(host='0.0.0.0', port=PORT, debug=True)
